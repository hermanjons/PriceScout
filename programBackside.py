# -*- coding: utf-8 -*-
"""
Created on Sun Sep 18 11:58:15 2022

@author: hamza IŞIK
"""

import sqlite3 as sql
from bs4 import BeautifulSoup#modÃ¼l dahil ediliyor
import datetime#dosyayÄ± tarihle kaydetmek iÃ§in modÃ¼l dahil ediliyor.
import pandas as pd
import numpy as np
import requests#modÃ¼l dahil ediliyor
from openpyxl import Workbook,load_workbook,styles
import openpyxl




def datefinder():
    an = datetime.datetime.now()
    tarih = datetime.datetime.strftime(an, '%d-%m-%Y')
    return tarih

def nameSelector():
    baglanti = sql.connect("veritabani.db")
    isaretci = baglanti.cursor()
    isaretci.execute("CREATE TABLE IF NOT EXISTS Dosya_sirasi (sira_no,kayit_tarih)")
    veri=isaretci.execute("SELECT sira_no FROM Dosya_sirasi")
    sayacListe=[]
    for i in veri:
        sayacListe.append(i[0])


    if len(sayacListe) < 1:
        sayac = 0
        isaretci.execute("INSERT INTO Dosya_sirasi VALUES ({},{})".format(sayac,datefinder()))
        baglanti.commit()
        return sayac
    else:

        finalsayac = max(sayacListe)
        finalsayac += 1
        isaretci.execute("INSERT INTO Dosya_sirasi VALUES ({},{})".format(finalsayac,datefinder()))
        baglanti.commit()
        return finalsayac




class ExcelReading:
    def __init__(self):
        self.stok_kod_liste=[]
        self.urun_ad_liste=[]
        self.excelrbcliste = []
        self.excelrbtstnliste = []
        self.excelrobolinkliste = []
        self.exceldirenclinkliste = []
        self.excelmotorobitliste = []
        self.excelrobitshopliste = []
        self.excelfile = []
        self.sayacrbc = 0
        self.sayacrbtshop = 0
        self.sayacrbtstn = 0
        self.sayacdirenc = 0
        self.sayacmtrobit = 0
        self.sayacrobolink = 0
        print("sayaclar oluşturuldu")


    def exceloku_kod(self):

        for beta in self.excelfile[0].get(0):
            try:
                self.stok_kod_liste.append(beta)
            except:
                self.stok_kod_liste.append("YOK")

    def exceloku_ad(self):

        for teta in self.excelfile[0].get(1):
            try:

                self.urun_ad_liste.append(teta)

            except:
                self.urun_ad_liste.append("YOK")
    def excelokurbc(self):

        for i in self.excelfile[0].get(2):
            try:


                self.excelrbcliste.append(i)


            except:

                self.excelrbcliste.append("YOK")


    def excelokurbtstn(self):

        for y in self.excelfile[0].get(3):
            try:
                self.excelrbtstnliste.append(y)
            except:
                self.excelrbtstnliste.append("YOK")

    def excelokurblink(self):


        for z in self.excelfile[0].get(4):
            try:
                self.excelrobolinkliste.append(z)

            except:

                self.excelrobolinkliste.append("YOK")

    def excelokudirenc(self):

        for t in self.excelfile[0].get(5):

            try:

                self.exceldirenclinkliste.append(t)

            except:
                self.exceldirenclinkliste.append("YOK")
                
    def excelokumtrobit(self):
        for m in self.excelfile[0].get(6):
            try:

                self.excelmotorobitliste.append(m)
            except:
                self.excelmotorobitliste.append("YOK")
    def excelokurbtshop(self):

        for s in self.excelfile[0].get(7):
            try:
                self.excelrobitshopliste.append(s)
            except:
                self.excelrobitshopliste.append("YOK")




class PriceFinder():

    def __init__(self,link=None,fpofTag=None,spofTag=None,tpofTag=None,
                 fsofTag=None,ssofTag=None,tsofTag=None):

        self.link=link
        self.fpofTag = fpofTag
        self.spofTag = spofTag
        self.tpofTag = tpofTag
        self.fsofTag = fsofTag
        self.ssofTag = ssofTag
        self.tsofTag = tsofTag



    def findpriceandstock(self):
        self.liste=[]
        self.getRequest = requests.get(self.link)
        self.source = BeautifulSoup(self.getRequest.content, "html.parser")


        try:

            self.stockdata = self.source.find("{}".format(self.fsofTag),
                                              attrs={"{}".format(self.ssofTag):"{}".format(self.tsofTag)})
            if len(self.stockdata) > 0:
                self.liste.insert(0,self.stockdata)
            else:
                pass

        except:
            self.liste.insert(0,"HATA")



        try:

            self.pricedata = self.source.find("{}".format(self.fpofTag),
                                              attrs={"{}".format(self.spofTag):"{}".format(self.tpofTag)})


            self.liste.insert(1, self.pricedata)

        except:
            self.liste.insert(1, "HATA")



        return self.liste[0],self.liste[1]








class PriceListing:

    def __init__(self):
        self.rbcfiyatlar = []
        self.rbckirmizi = []
        self.robotistanfiyatlar = []
        self.robotistankirmizi = []
        self.rblinkkirmizi = []
        self.robolinkfiyatlar= []
        self.direnckirmizi = []
        self.direncfiyatlar = []
        self.mtrobitfiyatlar = []
        self.mtrobitkirmizi = []
        self.robitshopfiyatlar = []
        self.robitshopkirmizi = []
        print("price listing sayacları olutşruuldu")



    def dataListRbc(self):

        for i in excelokuyucu.excelrbcliste:

            excelokuyucu.sayacrbc += 1
            if len(i) > 0:

                try:
                    datafinderrbc = PriceFinder(i,"span","class","right_line indirimliFiyat")
                    _,pricedatarbc = datafinderrbc.findpriceandstock()
                    pricedatarbc = pricedatarbc.find("span",attrs={"class":"spanFiyat"})
                    pricedatarbc=pricedatarbc.text.rstrip("₺")
                    self.rbcfiyatlar.append(pricedatarbc.strip())

                except:
                    
                    self.rbcfiyatlar.append("Link hata")

            else:
                self.rbcfiyatlar.append("LİNK GİR")


    def dataListRbtstn(self):

        for y in excelokuyucu.excelrbtstnliste:
            excelokuyucu.sayacrbtstn += 1

            if len(y) > 0:

                try:
                    fiyatbulurbtstn = PriceFinder(y, "span", "class", "product-price", "div", "class",
                                                     "w-100 p-1 out-stock-available", )
                    stockdatarbtstn, pricedatarbtstn = fiyatbulurbtstn.findpriceandstock()
                    pricedatarbtstn = pricedatarbtstn.text
                    print(pricedatarbtstn)
                    pricedatarbtstn = pricedatarbtstn.strip()
                    pricedatarbtstn = pricedatarbtstn.rstrip("TL")

                    try:

                        stockdatarbtstn = stockdatarbtstn.find("p", attrs={"class": "text-center fw-bold"})
                        stockdatarbtstn = stockdatarbtstn.text.strip()
                    except:
                        pass
                    if stockdatarbtstn == "Tükendi" and stockdatarbtstn != "HATA":
                        print("şart sağlandı")
                        self.robotistanfiyatlar.append(pricedatarbtstn.strip() + "-0")
                        self.robotistankirmizi.append(excelokuyucu.sayacrbtstn)

                    else:

                        self.robotistanfiyatlar.append(pricedatarbtstn.strip())

                except:
                    self.robotistanfiyatlar.append("link hata")

            else:
                self.robotistanfiyatlar.append("LİNK GİR")
                
    def dataListRobolink(self):


        for z in excelokuyucu.excelrobolinkliste:
            excelokuyucu.sayacrobolink += 1

            if len(z) > 0:

                try:
                    fiyatbulucurblnk = PriceFinder(z, "span", "class", "product-price", "div", "class",
                                                  "w-100 p-1 out-stock-available", )
                    stockdatarblnk, pricedatarblnk = fiyatbulucurblnk.findpriceandstock()
                    pricedatarblnk = pricedatarblnk.text

                    pricedatarblnk = pricedatarblnk.strip()
                    pricedatarblnk = pricedatarblnk.rstrip("TL")

                    try:

                        stockdatarblnk = stockdatarblnk.find("p", attrs={"class": "text-center fw-bold"})
                        stockdatarblnk = stockdatarblnk.text.strip()
                    except:
                        pass
                    if stockdatarblnk == "Tükendi" and stockdatarblnk != "HATA":
                        print("şart sağlandı")
                        self.robolinkfiyatlar.append(pricedatarblnk.strip() + "-0")
                        self.rblinkkirmizi.append(excelokuyucu.sayacrobolink)

                    else:

                        self.robolinkfiyatlar.append(pricedatarblnk.strip())

                except:
                    self.robolinkfiyatlar.append("link hata")

            else:
                self.robolinkfiyatlar.append("LİNK GİR")

    def dataListDirenc(self):

        for t in excelokuyucu.exceldirenclinkliste:
            excelokuyucu.sayacdirenc +=1
            if len(t) > 0 :


                try:
                    fiyatbulucudirenc = PriceFinder(t,"span","class","product-price-tl","span","class","box productFunction")
                    stockdatadirenc,pricedatadirenc = fiyatbulucudirenc.findpriceandstock()


                    if len(stockdatadirenc) > 0 and stockdatadirenc != "HATA":
                        self.direncfiyatlar.append(pricedatadirenc.text+"-0")
                        self.direnckirmizi.append(excelokuyucu.sayacdirenc)

                    else:
                        self.direncfiyatlar.append(pricedatadirenc.text)
                except:
                    self.direncfiyatlar.append("link hata")
            else:
                self.direncfiyatlar.append("LİNK GİR")
                
    def dataListMtRobit(self):
        

        for m in excelokuyucu.excelmotorobitliste:
            excelokuyucu.sayacmtrobit +=1
            
            if len(m) > 0:

                try:
                    fiyatbulucumtrobit = PriceFinder(m,"span","class","product-price-not-vat","div","class","w-100 p-1 out-stock-available",)
                    stockdatamtrobit,pricedatamtrobit = fiyatbulucumtrobit.findpriceandstock()
                    pricedatamtrobit = pricedatamtrobit.text
                    pricedatamtrobit = pricedatamtrobit.strip()
                    pricedatamtrobit=pricedatamtrobit.rstrip("TL")
                    try:

                        stockdatamtrobit = stockdatamtrobit.find("p",attrs={"class":"text-center fw-bold"})
                        stockdatamtrobit = stockdatamtrobit.text.strip()
                    except:
                        pass
                    if stockdatamtrobit == "Tükendi" and stockdatamtrobit != "HATA":
                        print("şart sağlandı")
                        self.mtrobitfiyatlar.append(pricedatamtrobit.strip()+"-0")
                        self.mtrobitkirmizi.append(excelokuyucu.sayacmtrobit)

                    else:

                        self.mtrobitfiyatlar.append(pricedatamtrobit.strip())

                except:
                    self.mtrobitfiyatlar.append("link hata")
                    
            else:
                self.mtrobitfiyatlar.append("LİNK GİR")

    def dataListRobitshop(self):


        for s in excelokuyucu.excelrobitshopliste:
            excelokuyucu.sayacrbtshop += 1

            if len(s) > 0:

                try:
                    veribulucurbtshop = PriceFinder(s,"div","class","product-price-old","div","class","product-buttons-row",)
                    stockdatarobitshp,pricedatarobitshop = veribulucurbtshop.findpriceandstock()
                    if len(stockdatarobitshp) > 0 and stockdatarobitshp.text.strip() == "Gelince Haber Ver":
                        self.robitshopfiyatlar.append(pricedatarobitshop.text.strip().rstrip("₺").strip()+"-0")
                        self.robitshopkirmizi.append(excelokuyucu.sayacrbtshop)

                    else:
                        self.robitshopfiyatlar.append(pricedatarobitshop.text.strip().rstrip("₺"))

                except Exception as e:

                    print(e)
                    self.robitshopfiyatlar.append("link hata")

            else:
                self.robitshopfiyatlar.append("LİNK GİR")


class cellColoring():
    def __init__(self,link):

        self.wb = load_workbook(link)
        self.ws = self.wb.active

    def redPaint(self):
        for i in fiyatliste.robotistankirmizi:
            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            self.ws["D{}".format(i + 1)].fill = my_fill
        for i in fiyatliste.rblinkkirmizi:
            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            self.ws["E{}".format(i + 1)].fill = my_fill
        for i in fiyatliste.direnckirmizi:
            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            self.ws["F{}".format(i + 1)].fill = my_fill
        for i in fiyatliste.mtrobitkirmizi:
            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            self.ws["G{}".format(i + 1)].fill = my_fill

        for i in fiyatliste.robitshopkirmizi:
            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
            self.ws["H{}".format(i + 1)].fill = my_fill



class ExcelCreator:



    def __init__(self,path):
        self.path = path



    def FinalExcelCreator(self):


        df = pd.DataFrame(list(zip(excelokuyucu.stok_kod_liste, excelokuyucu.urun_ad_liste, fiyatliste.rbcfiyatlar,
                               fiyatliste.robotistanfiyatlar,
                               fiyatliste.robolinkfiyatlar, fiyatliste.direncfiyatlar,
                               fiyatliste.mtrobitfiyatlar, fiyatliste.robitshopfiyatlar
                               )),
                      columns=['STOK KODU', 'ÜRÜN ADI', 'ROBOCOMBO', 'ROBOTİSTAN', 'ROBOLİNK', 'DİRENÇ', 'MOTOROBİT',
                               'ROBİTSHOP'])
        isim = "{}/{}-{}.xlsx".format(self.path,datefinder().rstrip("."), nameSelector())

        writer = pd.ExcelWriter(isim, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer._save()

        mycell = cellColoring(isim)
        mycell.redPaint()
        mycell.wb.save(isim)


excelokuyucu = ExcelReading()

fiyatliste = PriceListing()


